import streamlit as st
import pandas as pd
import datetime
from openpyxl import load_workbook
import os
import sys

# Conditional imports: Only import Windows-specific tools if on Windows
if sys.platform == "win32":
    import pythoncom
    import win32com.client as win32
else:
    pythoncom = None
    win32 = None

import PyPDF2
from playwright.sync_api import sync_playwright
import google.generativeai as genai
import zipfile
import io
import firebase_admin
from firebase_admin import credentials, firestore, storage

st.set_page_config(page_title="Forklift Training Portal", layout="wide", page_icon="🚜")

# --- DIRECTORY SETUP ---
# We no longer need the local license_photos folder!
os.makedirs("templates", exist_ok=True)

# --- FIREBASE SETUP ---
if not firebase_admin._apps:
    cert_dict = dict(st.secrets["firebase"])
    cred = credentials.Certificate(cert_dict)
    
    # Safely find the bucket name whether it is at the top of the file OR at the bottom
    bucket_name = st.secrets.get("FIREBASE_BUCKET", "")
    if not bucket_name and "firebase" in st.secrets:
        bucket_name = st.secrets["firebase"].get("FIREBASE_BUCKET", "")
        
    firebase_admin.initialize_app(cred, {
        'storageBucket': bucket_name
    })
db = firestore.client()
bucket = storage.bucket()

def load_data():
    docs = db.collection("employees").stream()
    data = []
    for doc in docs:
        emp_dict = doc.to_dict()
        emp_dict["Employee ID"] = doc.id 
        data.append(emp_dict)
        
    expected_columns = [
        "Employee ID", "Name", "Designation", "Cost Centre", 
        "Status", "Expiry Date", "Scheduled Date", 
        "Date of Birth", "Contact Number", "Email Address", "Manager Email"
    ]
    
    if data:
        df = pd.DataFrame(data)
        for col in expected_columns:
            if col not in df.columns:
                df[col] = ""
        df = df[expected_columns]
    else:
        df = pd.DataFrame(columns=expected_columns)
        
    df["Scheduled Date"] = df["Scheduled Date"].fillna("").astype(str).replace("Not Scheduled", "")
    df.loc[~df["Status"].isin(["Registration Sent", "Manager Approved"]), "Scheduled Date"] = ""
    return df

df = load_data()

# --- FORTIFIED AI DATA ANALYST MODULE ---
@st.cache_data(ttl=1800)
def get_dates_via_gemini(api_key):
    if not api_key:
        return ["AI Error: No API Key provided."]
    url = "https://www.ntuclearninghub.com/-/course/forklift-operator-refresher-training"
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(user_agent="Mozilla/5.0...")
            page = context.new_page()
            page.goto(url, wait_until="domcontentloaded")
            for i in range(5):
                page.evaluate("window.scrollBy(0, 800)")
                page.wait_for_timeout(1000) 
            page_text = page.locator("body").text_content() or ""
            page_text = " ".join(page_text.split())
            browser.close()
        
        genai.configure(api_key=api_key.strip())
        model = genai.GenerativeModel('gemini-2.5-flash')
        prompt = f"""
        You are an operational scheduling assistant. Carefully read the following scraped webpage text.
        Look closely for sections titled "Course Schedule", "Course Period", or "Course Dates".
        Extract the FULL training period (both start and end dates) for all upcoming sessions.
        Return ONLY a clean, comma-separated list of the dates found. 
        If the course is only one day long, just return the single date (for example: 19/06/2026).
        If the course spans multiple days, return it as "DD/MM/YYYY to DD/MM/YYYY".
        Do not include markdown, code blocks, or conversational text.
        If you absolutely cannot find any dates in the text, return exactly: "No dates found".
        Webpage Text: {page_text[:100000]} 
        """
        response = model.generate_content(prompt)
        raw_output = response.text.strip()
        if "No dates found" not in raw_output and raw_output:
            return [d.strip() for d in raw_output.split(',')]
        else:
            return ["No clear dates found on the website right now."]
    except Exception as e:
        if "429" in str(e):
            return ["API Error: Quota exceeded. Please check your Gemini account."]
        return [f"System Error: {e}"]

# ==========================================
# --- SMART LINK: MANAGER SCHEDULING VIEW ---
# ==========================================
query_params = st.query_params
if "manager" in query_params:
    mgr_email = query_params["manager"]
    st.title("🗓️ Manager Training Approval")
    st.markdown(f"Welcome. Please select a training schedule for the employees under **{mgr_email}**.")
    
    mgr_df = df[(df["Manager Email"] == mgr_email) & (df["Status"].str.lower().isin(["needs refresher", "need refresher"]))]
    
    if mgr_df.empty:
        st.success("All your employees are currently up to date or already scheduled!")
        st.stop()
        
    st.dataframe(mgr_df[["Employee ID", "Name", "Expiry Date"]], hide_index=True)
    
    st.markdown("### Select Date")
    api_key = st.secrets.get("GEMINI_API_KEY", "")
    if api_key:
        with st.spinner("Fetching live dates from NTUC..."):
            available_dates = get_dates_via_gemini(api_key)
        selected_date = st.selectbox("Select Training Schedule:", available_dates)
    else:
        selected_date = st.text_input("Enter Training Date (e.g., DD/MM/YYYY):")
    
    if st.button("✅ Confirm Schedule for these Employees", type="primary"):
        if selected_date and "Error" not in selected_date:
            for index, row in mgr_df.iterrows():
                emp_id = str(row["Employee ID"])
                db.collection("employees").document(emp_id).update({
                    "Status": "Manager Approved",
                    "Scheduled Date": selected_date
                })
            st.success("✅ Schedules confirmed! HR has been notified to send the registration to NTUC.")
            st.info("You may now close this window.")
        else:
            st.error("Please provide a valid date.")
    st.stop() 
# ==========================================

# ==========================================
# --- SECURITY: HR PASSWORD LOGIN SCREEN ---
# ==========================================
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    st.title("🔒 YCH Logistics Training Portal")
    st.markdown("Please enter the authorized team password to access the secure dashboard.")
    
    col1, col2 = st.columns([1, 2])
    with col1:
        entered_password = st.text_input("Portal Password", type="password")
        if st.button("Login", type="primary", use_container_width=True):
            if entered_password == st.secrets.get("APP_PASSWORD", "Team@ych2026"):
                st.session_state.authenticated = True
                st.rerun() 
            else:
                st.error("❌ Incorrect password. Please try again.")
    st.stop() 
# ==========================================


# --- HEADER ---
st.title("YCH Forklift Training & Refresher Portal")
st.markdown("Welcome! Use the tabs below to navigate through the dashboard, manage employees, schedule training, or process documents.")

# --- SIDEBAR (Global Settings & Backups Only) ---
st.sidebar.header("⚙️ App Settings")
st.sidebar.markdown("Configure your email and AI settings here.")

if "GEMINI_API_KEY" in st.secrets and st.secrets["GEMINI_API_KEY"]:
    gemini_api_key = st.secrets["GEMINI_API_KEY"]
    st.sidebar.success("✅ AI Connected")
else:
    gemini_api_key = st.sidebar.text_input("Gemini API Key", type="password")

sender_email = st.sidebar.text_input("Your Email (Sender)", value="kai.ayao@ych.com")
ntuc_email = st.sidebar.text_input("NTUC Contact Email", value="test@ntuclearninghub.com")

st.sidebar.divider()
st.sidebar.header("📥 Data Backups")
st.sidebar.markdown("Download a live copy of your database or photos from Google Cloud.")

def get_audited_database_csv():
    export_df = df.copy()
    if not export_df.empty:
        # Check cloud storage for existing photos
        blobs = bucket.list_blobs(prefix="license_photos/")
        cloud_files = [blob.name for blob in blobs]
        
        export_df["Old License File"] = export_df["Employee ID"].apply(lambda x: "In Cloud" if f"license_photos/{x}_old.jpg" in cloud_files else "Missing")
        export_df["New License File"] = export_df["Employee ID"].apply(lambda x: "In Cloud" if f"license_photos/{x}_new.jpg" in cloud_files else "Missing")
    return export_df.to_csv(index=False)

st.sidebar.download_button("📊 Download Live Database (CSV)", data=get_audited_database_csv(), file_name=f"forklift_database_{datetime.date.today()}.csv", mime="text/csv")

def create_photos_zip():
    zip_buffer = io.BytesIO()
    blobs = bucket.list_blobs(prefix="license_photos/")
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for blob in blobs:
            file_name = blob.name.split("/")[-1]
            if file_name: # Ensure it's not an empty folder blob
                # Download raw bytes from Google Cloud and put directly into the ZIP
                zip_file.writestr(file_name, blob.download_as_bytes())
    return zip_buffer.getvalue()

st.sidebar.download_button("🖼️ Download All Photos (ZIP)", data=create_photos_zip(), file_name=f"license_photos_{datetime.date.today()}.zip", mime="application/zip")


# ==========================================
# --- MAIN APP NAVIGATION (TABS) ---
# ==========================================
tab1, tab2, tab3, tab4 = st.tabs(["📊 Dashboard & Tracking", "👥 Manage Employees", "🎓 Training & Scheduling", "📁 Document Processing"])

# --- TAB 1: DASHBOARD ---
with tab1:
    st.header("Executive Overview")
    if not df.empty:
        dashboard_df = df.copy()
        dashboard_df["Status"] = dashboard_df["Status"].astype(str).str.strip().str.title()
        dashboard_df.loc[dashboard_df["Status"] == "Need Refresher", "Status"] = "Needs Refresher"
        
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total Operators", len(dashboard_df))
        
        needs_refresher_count = len(dashboard_df[dashboard_df["Status"] == "Needs Refresher"])
        m2.metric("Needs Refresher", needs_refresher_count, delta="-Action Req" if needs_refresher_count > 0 else None, delta_color="inverse")
        m3.metric("Registration Sent", len(dashboard_df[dashboard_df["Status"] == "Registration Sent"]))
        m4.metric("Completed", len(dashboard_df[dashboard_df["Status"] == "Completed"]))
        
        st.divider()
        col_chart, col_roster = st.columns([1, 2])
        
        with col_chart:
            st.markdown("##### Workforce Status")
            st.bar_chart(dashboard_df["Status"].value_counts(), color="#4CAF50")
            
        with col_roster:
            st.markdown("##### 🗓️ Upcoming Classes")
            roster_df = df[(df["Scheduled Date"].notna()) & (df["Scheduled Date"].str.strip() != "")]
            if not roster_df.empty:
                summary_data = [{"Training Date": date, "Total Pax": len(group), "Registered Employees": ", ".join(group["Name"].tolist())} for date, group in roster_df.groupby("Scheduled Date")]
                st.dataframe(pd.DataFrame(summary_data), hide_index=True, use_container_width=True)
            else:
                st.info("No upcoming classes scheduled.")

        st.divider()
        st.markdown("##### 🚨 License Expiry Tracker")
        st.caption("Rows highlighted in red require immediate attention (expiry < 30 days). Yellow indicates expiry within 90 days.")
        
        def highlight_expiry(row):
            status = str(row["Status"]).strip().lower()
            try:
                exp_date = pd.to_datetime(row["Expiry Date"], dayfirst=True).date()
                days_left = (exp_date - datetime.date.today()).days
                if days_left < 30 or status in ["needs refresher", "need refresher"]: return ['background-color: #ffcccc'] * len(row)
                elif days_left < 90: return ['background-color: #fff2cc'] * len(row)
            except:
                if status in ["needs refresher", "need refresher"]: return ['background-color: #ffcccc'] * len(row)
            return [''] * len(row)
            
        st.dataframe(df.style.apply(highlight_expiry, axis=1), use_container_width=True)
    else:
        st.info("No data available. Please add employees in the 'Manage Employees' tab.")

# --- TAB 2: MANAGE EMPLOYEES ---
with tab2:
    st.header("Employee Database Management")
    col_add, col_bulk = st.columns([2, 1])
    
    with col_add:
        st.subheader("Add or Update Single Employee")
        lookup_id = st.text_input("🔍 Enter Employee ID first (Press Enter to auto-fill)", help="Type an existing ID to update, or a new ID to create.")
        
        def_vals = {
            "Name": "", "Designation": "", "Cost Centre": "",
            "Email Address": "", "Contact Number": "", "Manager Email": "",
            "Status": "Needs Refresher", "Scheduled Date": ""
        }
        def_dob = datetime.date(1990, 1, 1)
        def_expiry = datetime.date.today()
        
        if lookup_id:
            existing = df[df["Employee ID"] == lookup_id]
            if not existing.empty:
                st.info(f"✅ Record found! Updating existing details for **{lookup_id}**.")
                rec = existing.iloc[0]
                for key in def_vals.keys():
                    if pd.notna(rec.get(key)) and str(rec.get(key)).strip() != "":
                        def_vals[key] = str(rec[key])
                try: def_dob = pd.to_datetime(rec["Date of Birth"], dayfirst=True).date()
                except: pass
                try: def_expiry = pd.to_datetime(rec["Expiry Date"], dayfirst=True).date()
                except: pass
            else:
                st.info("✨ New ID detected. Fill out the form below to add them.")
        
        # --- DYNAMIC COST CENTRE LOGIC ---
        st.markdown("##### 🏢 Department Details")
        
        # 1. Grab all unique cost centres currently in the database
        existing_ccs = sorted([str(cc).strip() for cc in df["Cost Centre"].unique() if str(cc).strip() != ""])
        cc_options = existing_ccs + ["Others (Add New)"]
        
        # 2. Figure out the default index
        current_cc = def_vals["Cost Centre"]
        if current_cc in existing_ccs:
            cc_index = existing_ccs.index(current_cc)
        elif current_cc: # Unlisted cost centre fallback
            cc_options.insert(0, current_cc)
            cc_index = 0
        else:
            cc_index = 0 if len(existing_ccs) > 0 else len(cc_options) - 1
            
        selected_cc = st.selectbox("Cost Centre", cc_options, index=cc_index)
        
        # 3. Show text input if 'Others' is selected
        if selected_cc == "Others (Add New)":
            final_cc = st.text_input("✏️ Type New Cost Centre Name:")
        else:
            final_cc = selected_cc

        st.markdown("##### 👤 Personal & License Details")
        with st.form("employee_form", clear_on_submit=False):
            
            # Row 1: Identity
            c1, c2 = st.columns(2)
            name = c1.text_input("Full Name *", value=def_vals["Name"])
            dob = c2.date_input("Date of Birth", value=def_dob, min_value=datetime.date(1940, 1, 1), max_value=datetime.date(2030, 12, 31))
            
            # Row 2: Contact
            c3, c4 = st.columns(2)
            contact_number = c3.text_input("Contact Number", value=def_vals["Contact Number"])
            email_address = c4.text_input("Employee Email", value=def_vals["Email Address"])
            
            # Row 3: Role & Management
            c5, c6 = st.columns(2)
            designation = c5.text_input("Designation", value=def_vals["Designation"])
            manager_email = c6.text_input("Manager Email", value=def_vals["Manager Email"])
            
            # Row 4: License & Status
            c7, c8 = st.columns(2)
            expiry_date = c7.date_input("License Expiry Date", value=def_expiry, min_value=datetime.date(1980, 1, 1), max_value=datetime.date(2050, 12, 31))
            
            status_options = ["Needs Refresher", "Manager Approved", "Registration Sent", "Completed"]
            try: stat_idx = status_options.index(def_vals["Status"])
            except: stat_idx = 0
            status = c8.selectbox("Status", status_options, index=stat_idx)
            
            st.divider()
            
            # Row 5: Extras
            c9, c10 = st.columns(2)
            scheduled_date = c9.text_input("Scheduled Date (Optional)", value=def_vals["Scheduled Date"], help="Only applies if Status is 'Registration Sent' or 'Manager Approved'")
            old_photo = c10.file_uploader("Upload Old License Photo (Optional)", type=["png", "jpg", "jpeg"])
            
            st.markdown("<br>", unsafe_allow_html=True) # Adds a little breathing room before the button
            
            if st.form_submit_button("💾 Save Employee Record", use_container_width=True):
                if lookup_id and name:
                    if not final_cc.strip():
                        st.error("⚠️ Please specify the Cost Centre.")
                    else:
                        # Upload photo directly to Google Cloud Storage
                        if old_photo:
                            blob = bucket.blob(f"license_photos/{lookup_id}_old.jpg")
                            blob.upload_from_string(old_photo.getvalue(), content_type=old_photo.type)
                            
                        if status not in ["Registration Sent", "Manager Approved"]: scheduled_date = ""
                        
                        new_record = {
                            "Name": name, "Designation": designation, "Cost Centre": final_cc.strip(),
                            "Date of Birth": str(dob), "Email Address": email_address, "Contact Number": contact_number,
                            "Manager Email": manager_email, "Expiry Date": str(expiry_date), "Status": status, "Scheduled Date": scheduled_date
                        }
                        db.collection("employees").document(lookup_id).set(new_record)
                        st.success(f"✅ Record saved successfully for {lookup_id}!")
                        st.rerun()
                else:
                    st.error("⚠️ Please enter an Employee ID and Full Name.")

    with col_bulk:
        st.subheader("Bulk Upload via CSV")
        st.markdown("Upload a filled CSV file to add or update multiple records at once.")
        template_df = pd.DataFrame(columns=["Employee ID", "Name", "Designation", "Cost Centre", "Date of Birth", "Email Address", "Contact Number", "Manager Email", "Expiry Date", "Status", "Scheduled Date"])
        st.download_button("📥 Download Empty Template", data=template_df.to_csv(index=False), file_name="employee_upload_template.csv", mime="text/csv", use_container_width=True)
        
        bulk_file = st.file_uploader("Upload Completed CSV", type=["csv"])
        if bulk_file and st.button("🚀 Process Bulk Upload", use_container_width=True):
            try:
                try: bulk_df = pd.read_csv(bulk_file)
                except UnicodeDecodeError:
                    bulk_file.seek(0)
                    bulk_df = pd.read_csv(bulk_file, encoding='latin1')
                if "Scheduled Date" not in bulk_df.columns: bulk_df["Scheduled Date"] = ""
                bulk_df = bulk_df.fillna("")
                
                with st.spinner("Uploading to Cloud Database..."):
                    for index, row in bulk_df.iterrows():
                        emp_id = str(row["Employee ID"])
                        record_dict = row.drop("Employee ID").to_dict()
                        db.collection("employees").document(emp_id).set(record_dict)
                st.success("✅ Bulk upload successful!")
                st.rerun() 
            except Exception as e:
                st.error(f"Error processing file: {e}")
        
        st.divider()
        st.subheader("📸 Cloud Photo Viewer")
        search_id = st.text_input("Enter Employee ID to view cloud license:")
        if search_id:
            # Pull photos directly from Google Cloud
            old_blob = bucket.blob(f"license_photos/{search_id}_old.jpg")
            new_blob = bucket.blob(f"license_photos/{search_id}_new.jpg")
            
            if old_blob.exists() or new_blob.exists():
                if old_blob.exists():
                    st.caption("Old License")
                    st.image(old_blob.download_as_bytes(), use_container_width=True)
                if new_blob.exists():
                    st.caption("Renewed License")
                    st.image(new_blob.download_as_bytes(), use_container_width=True)
            else:
                st.warning("No photos found in Google Cloud for this ID.")

# --- TAB 3: TRAINING & SCHEDULING ---
with tab3:
    st.header("Automated Scheduling & Communications")
    LIVE_APP_URL = "http://10.202.53.52" 
    
    col_mgr, col_ntuc = st.columns(2)
    with col_mgr:
        st.subheader("1. 📢 Request Manager Approvals")
        st.markdown("Send an email to managers with a smart link to pick dates for their staff.")
        
        needs_refresher_mask = df["Status"].str.lower().isin(["needs refresher", "need refresher"])
        needs_refresher_list = df[needs_refresher_mask]
        
        if not needs_refresher_list.empty:
            valid_managers = [m for m in needs_refresher_list["Manager Email"].unique() if pd.notna(m) and str(m).strip() != ""]
            if valid_managers:
                for mgr_email in valid_managers:
                    mgr_emps = needs_refresher_list[needs_refresher_list["Manager Email"] == mgr_email]
                    if st.button(f"📧 Draft Link Email to {mgr_email}", use_container_width=True):
    if sys.platform == "win32":
        try:
            smart_link = f"{LIVE_APP_URL}/?manager={mgr_email}"
            pythoncom.CoInitialize()
            outlook = win32.Dispatch('outlook.application')
            # ... (the rest of your mail code) ...
            mail.Display()
            st.success("✅ Email drafted!")
        except Exception as e:
            st.error(f"⚠️ Error: {e}")
    else:
        st.error("⚠️ Outlook automation only works on Windows office computers.")                            
                            html_table = "<table style='border-collapse: collapse; width: 100%;'><tr style='background-color: #f2f2f2;'><th style='border: 1px solid #ddd; padding: 8px;'>Name</th><th style='border: 1px solid #ddd; padding: 8px;'>ID</th><th style='border: 1px solid #ddd; padding: 8px;'>Expiry Date</th></tr>"
                            for _, row in mgr_emps.iterrows():
                                html_table += f"<tr><td style='border: 1px solid #ddd; padding: 8px;'>{row['Name']}</td><td style='border: 1px solid #ddd; padding: 8px;'>{row['Employee ID']}</td><td style='border: 1px solid #ddd; padding: 8px; color: red;'>{row['Expiry Date']}</td></tr>"
                            html_table += "</table>"
                            
                            mail.HTMLBody = f"""
                            <p>Hi,</p>
                            <p>The following employees under your supervision are due for mandatory Forklift Refresher Training:</p>
                            {html_table}
                            <br>
                            <p><b><a href='{smart_link}' style='font-size: 16px; color: #ffffff; background-color: #0052cc; padding: 10px 20px; text-decoration: none; border-radius: 5px;'>CLICK HERE TO SELECT THEIR TRAINING DATES</a></b></p>
                            <br>
                            <p>Best regards,<br>{sender_email}</p>
                            """
                            mail.Display()
                            st.success(f"✅ Email drafted for {mgr_email}!")
                        except Exception as e:
                            st.error(f"⚠️ Connection error: {e}")
            
            missing_managers = needs_refresher_list[~needs_refresher_list["Manager Email"].isin(valid_managers)]
            if not missing_managers.empty:
                st.warning("⚠️ The following employees have no Manager Email assigned:")
                st.dataframe(missing_managers[["Name", "Employee ID"]], hide_index=True)
        else:
            st.success("Great news! No employees currently require schedule requests.")

    with col_ntuc:
        st.subheader("2. ✉️ Draft NTUC Registrations")
        st.markdown("Draft the final email to NTUC for employees who have been approved by their managers.")
        
        approved_mask = df["Status"] == "Manager Approved"
        approved_list = df[approved_mask]
        
        if not approved_list.empty:
            st.warning(f"🚨 You have {len(approved_list)} employees approved by managers waiting to be sent to NTUC!")
            st.dataframe(approved_list[["Name", "Scheduled Date", "Manager Email"]], hide_index=True)
            
            if st.button("✉️ Draft Bulk NTUC Email & Mark as 'Registration Sent'", type="primary", use_container_width=True):
                output_filename = f"NTUC_Bulk_Registration_{datetime.date.today()}.xlsx"
                try:
                    wb = load_workbook("templates/ntuc_template.xlsx")
                    ws = wb.active
                    def write_cell(sheet, coord, val):
                        cell = sheet[coord]
                        if type(cell).__name__ == 'MergedCell':
                            for mr in sheet.merged_cells.ranges:
                                if coord in mr:
                                    sheet.cell(row=mr.min_row, column=mr.min_col).value = val
                                    return
                        else: cell.value = val

                    primary_date = approved_list.iloc[0]["Scheduled Date"]
                    parts = str(primary_date).split(" to ") if " to " in str(primary_date) else [str(primary_date), str(primary_date)]
                    write_cell(ws, "J4", parts[0].strip())
                    write_cell(ws, "L4", parts[-1].strip())
                    
                    for idx, row in approved_list.reset_index().iterrows():
                        row_num = 12 + idx
                        write_cell(ws, f"A{row_num}", str(idx + 1))                     
                        write_cell(ws, f"B{row_num}", row["Name"])        
                        write_cell(ws, f"C{row_num}", "Forklift Operator")     
                        write_cell(ws, f"E{row_num}", row["Employee ID"]) 
                        try: formatted_dob = pd.to_datetime(str(row["Date of Birth"]), dayfirst=True).strftime("%d/%m/%Y")
                        except: formatted_dob = str(row["Date of Birth"])
                        write_cell(ws, f"G{row_num}", formatted_dob) 
                    wb.save(output_filename)

                    pythoncom.CoInitialize()
                    import win32com.client as win32
                    outlook = win32.Dispatch('outlook.application')
                    mail = outlook.CreateItem(0)
                    mail.SentOnBehalfOfName = sender_email
                    mail.To = ntuc_email
                    mail.Subject = f"Forklift Refresher Bulk Registration - {len(approved_list)} Pax"
                    mail.Body = f"Hi NTUC Team,\n\nPlease find attached the bulk refresher course registration for {len(approved_list)} employees.\n\nThank you,\n{sender_email}"
                    mail.Attachments.Add(os.path.abspath(output_filename))
                    mail.Display()
                    
                    for emp_id in approved_list["Employee ID"]:
                        db.collection("employees").document(str(emp_id)).update({"Status": "Registration Sent"})
                    st.success(f"✅ Email drafted! Database updated to 'Registration Sent'.")
                    st.rerun()
                except Exception as e:
                    st.error(f"⚠️ Error: {e}")
        else:
            st.info("Waiting for managers to approve schedules via the email links.")

# --- TAB 4: DOCUMENT PROCESSING ---
with tab4:
    st.header("Document & Certificate Processing")
    st.markdown("Automate status updates and securely store license documentation.")
    
    col_pdf, col_photo = st.columns(2)
    with col_pdf:
        st.subheader("📄 Auto-Scan NTUC Certificates")
        st.markdown("Upload NTUC PDF invoices/certificates. The system scans the text to find Employee IDs and automatically updates their status to **Completed**.")
        uploaded_pdfs = st.file_uploader("Drop NTUC PDFs here", type=["pdf"], accept_multiple_files=True)
        
        if uploaded_pdfs and st.button("🔍 Scan PDFs & Update Database", type="primary", use_container_width=True):
            updated_employees = []
            with st.spinner("Scanning documents using OCR..."):
                for pdf_file in uploaded_pdfs:
                    try:
                        reader = PyPDF2.PdfReader(pdf_file)
                        extracted_text = "".join([page.extract_text() for page in reader.pages])
                        for index, row in df.iterrows():
                            emp_name, emp_id = str(row["Name"]), str(row["Employee ID"])
                            if emp_id in extracted_text or emp_name.upper() in extracted_text.upper():
                                if df.at[index, "Status"] != "Completed":
                                    db.collection("employees").document(emp_id).update({"Status": "Completed", "Scheduled Date": ""})
                                    updated_employees.append(f"{emp_name} ({emp_id})")
                    except Exception as e:
                        st.error(f"Could not read {pdf_file.name}. Error: {e}")
            
            if updated_employees:
                st.success(f"✅ Successfully updated **{len(updated_employees)}** records to 'Completed'!")
                st.rerun()
            else:
                st.info("No matching employees found in the uploaded documents.")

    with col_photo:
        st.subheader("📸 Upload Renewed License Photos")
        st.markdown("Attach the final physical license photo to an employee's file directly into Google Cloud once they have completed training.")
        
        completed_emps = []
        if not df.empty and "Status" in df.columns:
            # Grab all employees marked as completed
            completed_emps = df[df["Status"] == "Completed"]["Employee ID"].tolist()

        if completed_emps:
            # Fetch a list of all photos currently sitting in Google Cloud
            blobs = bucket.list_blobs(prefix="license_photos/")
            cloud_files = [blob.name for blob in blobs]
            
            # Filter the list: Only keep IDs that DO NOT have a _new.jpg file in the cloud
            pending_photo_emps = [
                emp for emp in completed_emps 
                if f"license_photos/{emp}_new.jpg" not in cloud_files
            ]
            
            # Check if there is anyone left in the pending list
            if pending_photo_emps:
                with st.form("new_photo_form", clear_on_submit=True):
                    selected_emp_photo = st.selectbox("Select Employee (Missing Photo)", pending_photo_emps)
                    new_photo = st.file_uploader("Upload New License Image", type=["png", "jpg", "jpeg"])
                    
                    if st.form_submit_button("☁️ Save to Google Cloud", use_container_width=True) and new_photo:
                        # Upload photo directly to Google Cloud Storage
                        blob = bucket.blob(f"license_photos/{selected_emp_photo}_new.jpg")
                        blob.upload_from_string(new_photo.getvalue(), content_type=new_photo.type)
                        
                        st.success(f"✅ Renewed photo saved to the cloud for {selected_emp_photo}!")
                        st.rerun()
            else:
                st.success("🎉 All 'Completed' employees currently have their new photos uploaded!")
        else:
            st.info("💡 You must update an employee's status to 'Completed' before you can attach their new license photo here.")
